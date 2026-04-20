"""core/import_parsers/solarwinds_parser.py — SolarWinds CSV/XLSX import.

SolarWinds doesn't have a single canonical export format; admins
typically run a SWQL query and save the result to CSV or Excel. So
the import is *column-map driven* — the user tells us which source
column maps to which PingWatch field on upload.

Two-step API:

    inspect_solarwinds(payload, filename)
      → {headers: [...], sample_rows: [first 5 rows], detected_format}
      Used by the column-mapper UI to show the user which fields to map.

    parse_solarwinds(payload, column_map, fmt)
      → canonical parse result (devices/errors/mapping_report)
      `column_map` keys are source headers; values are PingWatch field
      names from `_VALID_TARGETS` (or "(skip)" / missing → ignored).

Every imported device gets a default `ping` sensor — the SolarWinds
schema rarely carries enough info to pick richer sensors. Admins can
remove the ping sensor in the review screen, or add others manually.

XLSX uses openpyxl (added to requirements.txt). When the dep isn't
installed, the parser returns a clean row-0 error rather than crashing;
the route layer translates it to a 503 "openpyxl not installed".
"""

from __future__ import annotations

import csv
import io

# Headers we know how to map from a source column to a PingWatch field.
_VALID_TARGETS = {
    "name", "host", "group", "external_id",
    "snmp_community_default", "snmp_version_default", "webhook_url",
}

# Sentinel value used by the column-mapper UI for "this column is ignored".
_SKIP_SENTINELS = {"", "(skip)", "skip", "ignore", "(ignore)", None}


# ── Inspection: produce headers + sample for the column-mapper UI ──

def inspect_solarwinds(payload, filename: str = "") -> dict:
    """Probe a SW upload to extract headers + a 5-row sample.

    `payload` is `str` for CSV, `bytes` for XLSX. `filename` (or its
    extension) chooses the path. Returns:

        {headers: [...], sample_rows: [[...], ...], detected_format: "csv"|"xlsx",
         error: str | None}

    On failure, `error` is a human-readable string and the other fields
    are empty.
    """
    fmt = _detect_format(filename, payload)
    if fmt == "xlsx":
        return _inspect_xlsx(payload)
    return _inspect_csv(payload)


def parse_solarwinds(payload, column_map: dict | None,
                     fmt: str | None = None,
                     filename: str = "") -> dict:
    """Parse a SW CSV/XLSX with the user-supplied column map.

    `column_map`: `{source_header: pingwatch_field}`. Required mappings
    are `name` and `host` — the parser fails fast if either is missing
    from the map (so the UI never gets a confusing per-row error).

    Returns canonical parse result.
    """
    column_map = _normalize_column_map(column_map or {})
    targets = set(column_map.values())
    missing = [t for t in ("name", "host") if t not in targets]
    if missing:
        return _empty_result(
            f"column_map missing required target(s): {', '.join(missing)}"
        )
    actual_fmt = fmt or _detect_format(filename, payload)
    if actual_fmt == "xlsx":
        rows = _read_xlsx_rows(payload)
        if isinstance(rows, dict):     # error dict
            return _empty_result(rows["error"])
    else:
        rows = _read_csv_rows(payload)
        if isinstance(rows, dict):
            return _empty_result(rows["error"])

    if not rows:
        return _empty_result("no data rows")

    devices: list = []
    errors:  list = []
    sensors_total  = 0   # one ping per device
    sensors_mapped = 0

    for idx, row in enumerate(rows):
        # Map the source row to PingWatch fields.
        rec: dict = {}
        for src_col, target in column_map.items():
            if target in _SKIP_SENTINELS or target not in _VALID_TARGETS:
                continue
            v = row.get(src_col)
            if v is not None and str(v).strip():
                rec[target] = str(v).strip()

        name = rec.get("name", "").strip()
        host = rec.get("host", "").strip()
        if not name:
            errors.append({"row": idx, "reason": "missing required field: name"})
            continue
        if not host:
            errors.append({"row": idx, "reason": "missing required field: host"})
            continue

        ext_id = rec.get("external_id") or f"sw:row{idx + 1}"
        if not ext_id.startswith("sw:"):
            ext_id = f"sw:{ext_id}"

        dev: dict = {
            "external_id": ext_id,
            "name":  name,
            "host":  host,
            "group": rec.get("group", ""),
            "sensors": [{"name": "PING", "stype": "ping"}],
        }
        for fld in ("webhook_url", "snmp_community_default",
                    "snmp_version_default"):
            v = rec.get(fld)
            if v:
                dev[fld] = v
        devices.append(dev)
        sensors_total  += 1
        sensors_mapped += 1

    return {
        "devices": devices,
        "errors":  errors,
        "mapping_report": {
            "sensors_total":   sensors_total,
            "sensors_mapped":  sensors_mapped,
            "sensors_skipped": [],
        },
    }


# ── Format detection + readers ─────────────────────────────────────

def _detect_format(filename: str, payload) -> str:
    """Return 'xlsx' or 'csv'. Filename extension wins; falls back to
    sniffing the payload (xlsx files start with PK zip magic).
    """
    fn = (filename or "").lower()
    if fn.endswith(".xlsx") or fn.endswith(".xlsm"):
        return "xlsx"
    if fn.endswith(".csv") or fn.endswith(".txt"):
        return "csv"
    if isinstance(payload, (bytes, bytearray)) and payload[:2] == b"PK":
        return "xlsx"
    return "csv"


def _inspect_csv(payload) -> dict:
    text = _coerce_text(payload)
    if text.startswith("\ufeff"):
        text = text[1:]
    if not text.strip():
        return {"headers": [], "sample_rows": [],
                "detected_format": "csv",
                "error": "empty file"}
    try:
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
    except Exception as e:
        return {"headers": [], "sample_rows": [],
                "detected_format": "csv",
                "error": f"CSV parse failed: {e}"}
    if not rows:
        return {"headers": [], "sample_rows": [],
                "detected_format": "csv",
                "error": "no rows"}
    headers = [(h or "").strip() for h in rows[0]]
    sample  = [_pad_row(r, len(headers)) for r in rows[1:6]]
    return {"headers": headers, "sample_rows": sample,
            "detected_format": "csv", "error": None}


def _inspect_xlsx(payload) -> dict:
    rows_or_err = _read_xlsx_rows(payload, max_rows=6)
    if isinstance(rows_or_err, dict):
        return {"headers": [], "sample_rows": [],
                "detected_format": "xlsx",
                "error": rows_or_err["error"]}
    if not rows_or_err:
        return {"headers": [], "sample_rows": [],
                "detected_format": "xlsx",
                "error": "no rows"}
    # `_read_xlsx_rows` returns dict-rows; reconstruct headers from first row.
    # Headers were stashed in row order; recover by reading raw again.
    headers, sample = _xlsx_headers_and_sample(payload)
    return {"headers": headers, "sample_rows": sample,
            "detected_format": "xlsx", "error": None}


def _read_csv_rows(payload):
    text = _coerce_text(payload)
    if text.startswith("\ufeff"):
        text = text[1:]
    if not text.strip():
        return {"error": "empty file"}
    try:
        reader = csv.DictReader(io.StringIO(text))
        if reader.fieldnames is None:
            return {"error": "no header row"}
        # Strip whitespace from headers — common in SWQL exports.
        reader.fieldnames = [(h or "").strip() for h in reader.fieldnames]
        return list(reader)
    except Exception as e:
        return {"error": f"CSV parse failed: {e}"}


def _read_xlsx_rows(payload, max_rows: int | None = None):
    """Read XLSX rows as dicts (header-keyed). Returns list-of-dicts,
    or `{"error": str}` on failure.
    """
    try:
        import openpyxl  # type: ignore
    except ImportError:
        return {"error": "openpyxl not installed — XLSX import unavailable. "
                         "Run the setup wizard or pip install openpyxl."}
    if not isinstance(payload, (bytes, bytearray)):
        return {"error": "XLSX payload must be bytes"}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(payload), read_only=True,
                                     data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return {"error": "empty workbook"}
        headers = [(_xlsx_cell(h) or "").strip() for h in header_row]
        out: list = []
        for i, raw in enumerate(rows_iter):
            if max_rows is not None and i >= max_rows:
                break
            rec = {}
            for j, val in enumerate(raw):
                if j >= len(headers):
                    break
                key = headers[j] or f"col{j+1}"
                rec[key] = _xlsx_cell(val)
            out.append(rec)
        return out
    except Exception as e:
        return {"error": f"XLSX parse failed: {e}"}


def _xlsx_headers_and_sample(payload, sample_size: int = 5
                              ) -> tuple[list, list]:
    """Return `(headers, sample_rows_as_lists)`. Used by inspect path."""
    try:
        import openpyxl  # type: ignore
    except ImportError:
        return [], []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(payload), read_only=True,
                                     data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return [], []
        headers = [(_xlsx_cell(h) or "").strip() for h in header_row]
        sample: list = []
        for i, raw in enumerate(rows_iter):
            if i >= sample_size:
                break
            sample.append([_xlsx_cell(v) for v in raw[:len(headers)]])
        return headers, sample
    except Exception:
        return [], []


def _xlsx_cell(v):
    """Normalize an openpyxl cell value to a JSON-serializable scalar."""
    if v is None:
        return ""
    if isinstance(v, (int, float, bool, str)):
        return v
    return str(v)


def _normalize_column_map(m: dict) -> dict:
    """Strip whitespace + drop entries with skip-sentinel targets."""
    out: dict = {}
    for src, dst in (m or {}).items():
        if src is None or dst in _SKIP_SENTINELS:
            continue
        out[str(src).strip()] = str(dst).strip()
    return out


def _coerce_text(payload) -> str:
    if isinstance(payload, str):
        return payload
    if isinstance(payload, (bytes, bytearray)):
        try:
            return payload.decode("utf-8")
        except UnicodeDecodeError:
            return payload.decode("latin-1", errors="replace")
    return ""


def _pad_row(row, width: int) -> list:
    row = list(row)
    if len(row) < width:
        row += [""] * (width - len(row))
    return row[:width]


def _empty_result(reason: str) -> dict:
    return {
        "devices": [],
        "errors":  [{"row": 0, "reason": reason}],
        "mapping_report": {
            "sensors_total":   0,
            "sensors_mapped":  0,
            "sensors_skipped": [],
        },
    }
